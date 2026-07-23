import streamlit as st
import pandas as pd
import numpy as np
import random
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# ==========================================
# 0. 頂層小幫手函數
# ==========================================
def get_team_of(staff_name, df_daily):
    ts = df_daily[df_daily['姓名'] == staff_name]['組別'].values
    return str(ts[0]).strip().upper() if len(ts) > 0 else ""

def get_team_priority(staff_name, df_daily):
    t = get_team_of(staff_name, df_daily)
    if t == 'G': return 1
    elif t == 'F': return 2
    elif t == 'E': return 3
    elif t == 'D': return 4
    elif t == 'C': return 5
    elif t == 'B': return 6
    elif t == 'A': return 7
    elif t == 'H': return 9 # H組為影子
    else: return 8

def get_zone_count(z, m_counts, s_name):
    if z in ['A2', 'B2', 'C2']: return sum(m_counts[s_name].get(x, 0) for x in ['A2', 'B2', 'C2'])
    if z in ['MO', 'MO1', 'MO2']: return sum(m_counts[s_name].get(x, 0) for x in ['MO', 'MO1', 'MO2'])
    if z in ['R1', 'R3']: return sum(m_counts[s_name].get(x, 0) for x in ['R1', 'R3'])
    if z in ['GB', 'GC']: return sum(m_counts[s_name].get(x, 0) for x in ['GB', 'GC'])
    return m_counts[s_name].get(z, 0)

def get_macro(zone):
    if zone in ["MO", "MO1", "MO2"]: return "OBS"
    if zone in ["S", "S1", "S2"]: return "SURG"
    if zone in ["R", "R1", "R2", "R3"]: return "RESUS"
    if zone in ["A1", "A2", "B1", "B2", "C1", "C2"]: return "MED"
    if zone in ["GB", "GC", "GD"]: return "G"
    if zone in ["T", "T2"]: return "TRIAGE"
    if zone in ["P", "P2"]: return "PEDS"
    return zone

def is_severe(zone):
    return zone in ['R', 'S', 'C2']

# ==========================================
# 🌟 AI 權重計分系統 (第9步驟核心)
# ==================================================
def get_zone_score(zone, name, team, day_idx, df_result, date_columns, monthly_counts, work_blocks, current_day_macro_teams, ab_gb_gc_fulfilled):
    
    # 9.4 均勻分配：月累積次數越少，分數越低 (優先級越高)
    score = get_zone_count(zone, monthly_counts, name) * 10000 

    past_zones = []
    for i in [1, 2]: 
        if day_idx - i >= 0:
            pz = str(df_result.loc[df_result['姓名'] == name, date_columns[day_idx - i]].values[0]).strip()
            if pz not in ['OFF', '', 'X', 'NAN', 'L', 'L2', '職代', 'None', '⚠️需指派老師']:
                past_zones.append(pz)

    past_1_zone = past_zones[0] if len(past_zones) > 0 else ""
    past_macros = [get_macro(pz) for pz in past_zones]

    # A組特權防撞
    if team == 'A' and zone in ['R2', 'A2', 'S1']:
        score += 5000000 

    # 9.1 嚴格避開「同大區、同組別」防撞
    macro_z = get_macro(zone)
    if team in current_day_macro_teams.get(macro_z, set()):
        score += 20000 

    # 9.2 嚴格避開 R, B1, C2, S 任何兩個連續安排
    severe_zones = ['R', 'B1', 'C2', 'S']
    if zone in severe_zones and past_1_zone in severe_zones:
        score += 10000000 

    # T與P 不超過3天防線 (盡量小於4天)
    if zone in ['T', 'T2', 'P', 'P2']:
        if past_1_zone not in ['T', 'T2', 'P', 'P2']: 
            days_to_off = work_blocks[name][day_idx]
            if days_to_off > 3: # 距離OFF超過3天，禁止開啟T/P副本
                score += 5000000 

    # 9.3 GB, GC 盡可能平均分配給組別 A, B, C
    if zone in ['GB', 'GC']:
        if team not in ['A', 'B', 'C']:
            score += 5000000 # G, F, E, D 嚴禁去 GB, GC
        else:
            c = get_zone_count(zone, monthly_counts, name)
            score += (c * 20000) # 給 A, B, C 更強的次數平衡引力

    # A2, B2, C2 天花板
    if zone in ['A2', 'B2', 'C2']:
        c = get_zone_count(zone, monthly_counts, name)
        if c >= 5: score += 5000000

    # MO系 遞增阻力
    if zone in ['MO', 'MO1', 'MO2']:
        c = get_zone_count(zone, monthly_counts, name)
        score += (c * 25000) 
        if c >= 5: score += 2000000 

    # 區域優先引力
    if zone == 'B1':
        if team == 'B': score -= 400
        elif team == 'A': score -= 300
        elif team == 'C': score -= 200
    elif zone == 'C1':
        if team == 'E': score -= 500
        elif team == 'D': score -= 400
        elif team == 'C': score -= 300
        elif team == 'B': score -= 200
        elif team == 'A': score -= 100
    elif zone == 'R' or zone == 'S':
        if team == 'A': score -= 400
        elif team == 'B': score -= 300
        elif team == 'C': score -= 200
    elif zone in ['A2', 'B2', 'C2']:
        if team == 'D': score -= 400
        elif team == 'C': score -= 300
        elif team == 'B': score -= 200
        elif team == 'A': score -= 100

    return score

# ==========================================
# 網頁 UI 初始化
# ==========================================
st.set_page_config(page_title="自動排班系統", layout="wide")
st.title("急診護理人員自動初步排班系統")
st.markdown("---")

col1, col2 = st.columns(2)
with col1: training_file = st.file_uploader("📂 1. 上傳班表", type=['xlsx', 'csv'])
with col2: template_file = st.file_uploader("📂 2. 上傳空白檔", type=['xlsx', 'csv'])

all_staff, df_shift, df_template, data_ready, date_columns = [], None, None, False, []

if training_file and template_file:
    try:
        df_shift = pd.read_csv(training_file) if 'csv' in training_file.name.lower() else pd.read_excel(training_file)
        df_template = pd.read_csv(template_file) if 'csv' in template_file.name.lower() else pd.read_excel(template_file)
            
        shift_cols = list(df_shift.columns)
        shift_cols[0], shift_cols[1], shift_cols[2], shift_cols[3] = '編號', '組別', '性別', '姓名'
        date_length = len(shift_cols) - 4
        shift_cols[4:] = [str(i) for i in range(1, date_length + 1)]
        df_shift.columns = shift_cols
        
        template_cols = list(df_template.columns)
        template_cols[0], template_cols[1], template_cols[2], template_cols[3] = '編號', '組別', '性別', '姓名'
        template_cols[4:4+date_length] = [str(i) for i in range(1, date_length + 1)]
        df_template.columns = template_cols
        
        df_shift, df_template = df_shift.dropna(subset=['姓名']), df_template.dropna(subset=['姓名'])
        df_shift['姓名'], df_template['姓名'] = df_shift['姓名'].astype(str).str.strip(), df_template['姓名'].astype(str).str.strip()
        all_staff, date_columns, data_ready = df_shift['姓名'].unique().tolist(), df_shift.columns[4:], True
        st.success("✅ 檔案載入成功！")
    except Exception as e:
        st.error(f"檔案讀取失敗：{e}")

# ==========================================
# 左側設定區 
# ==========================================
st.sidebar.header("⚙️ 本月特殊排班規則設定")
train_s2 = st.sidebar.multiselect("S2 訓練名單", options=all_staff if data_ready else [])
train_b1_r = st.sidebar.multiselect("B1/R/C2/S 訓練名單", options=all_staff if data_ready else [])
train_t = st.sidebar.multiselect("檢傷(T) 訓練名單", options=all_staff if data_ready else [])

leader_options = ["請點選"] + all_staff if data_ready else ["請點選"]

st.sidebar.subheader("👑 各班組長順位")
d_l_1 = st.sidebar.selectbox("D班 第一順位", leader_options)
d_l_2 = st.sidebar.selectbox("D班 第二順位", leader_options)
d_l_3 = st.sidebar.selectbox("D班 第三順位", leader_options)
e_l_1 = st.sidebar.selectbox("E班 第一順位", leader_options)
e_l_2 = st.sidebar.selectbox("E班 第二順位", leader_options)
e_l_3 = st.sidebar.selectbox("E班 第三順位", leader_options)
n_l_1 = st.sidebar.selectbox("N班 第一順位", leader_options)
n_l_2 = st.sidebar.selectbox("N班 第二順位", leader_options)
n_l_3 = st.sidebar.selectbox("N班 第三順位", leader_options)

master_zones = ["T", "A1", "B1", "C1", "A2", "B2", "C2", "R", "R1", "R2", "R3", "P", "MO", "MO1", "MO2", "S1", "S", "S2", "T2", "GB", "GC", "GD"]

st.sidebar.markdown("---")
st.sidebar.subheader("🌱 H組 新人臨床教師與區域指定")

h_team_members = []
h_team_config = {}

if data_ready:
    h_team_members = df_shift[df_shift['組別'].astype(str).str.strip().str.upper() == 'H']['姓名'].unique().tolist()
    
    if not h_team_members:
        st.sidebar.info("✅ 本月班表中未偵測到「H組」未獨立新人。")
    else:
        st.sidebar.success(f"系統自動抓取到 {len(h_team_members)} 位 H組 新人！")
        for i, nh_name in enumerate(h_team_members):
            with st.sidebar.expander(f"👤 {nh_name} 專屬設定", expanded=True):
                p1 = st.selectbox(f"第一順位教師", ["無"] + all_staff, key=f"p1_{i}")
                p2 = st.selectbox(f"第二順位教師", ["無"] + all_staff, key=f"p2_{i}")
                p3 = st.selectbox(f"第三順位教師", ["無"] + all_staff, key=f"p3_{i}")
                
                allowed = st.multiselect(
                    f"可安排區域 (放假後換區用)", 
                    options=master_zones, 
                    default=["MO", "MO1", "MO2", "A1", "B1", "C1"],
                    key=f"allowed_{i}"
                )
                
                h_team_config[nh_name] = {
                    "preceptors": [p for p in [p1, p2, p3] if p != "無"],
                    "allowed_zones": allowed
                }

st.markdown("---")
main_col, help_col = st.columns([7.5, 2.5])

with help_col:
    st.markdown("### 💡 常見錯誤排除指南")
    st.info("""
    如果你按下排班按鈕後發生 **紅色錯誤 (Error)** 或跑不出結果，請檢查以下 4 點：
    **1️⃣ 檢查 Excel 表頭格式是否正確？** (編號、組別、性別、姓名)
    **2️⃣ 兩份 Excel 的「姓名」與「天數」是否吻合？**
    **3️⃣ 左側「各班組長順位」選了嗎？**
    **4️⃣ 檢查特殊區域代號是否填錯？**
    """)

with main_col:
    if st.button("🚀 開始自動排班運算", disabled=not data_ready, use_container_width=True):
        with st.spinner("🚀 引擎啟動，正在執行 9 大排班護盤步驟..."):
            try:
                df_result = df_template.copy()
                monthly_counts = {name: {} for name in all_staff}
                progress_bar = st.progress(0)
                missing_l_records = []
                
                work_blocks = {name: [0]*len(date_columns) for name in all_staff}
                for name in all_staff:
                    for day_idx in range(len(date_columns)):
                        count = 0
                        for future_idx in range(day_idx, len(date_columns)):
                            col = date_columns[future_idx]
                            val = str(df_shift.loc[df_shift['姓名'] == name, col].values[0]).strip().upper()
                            if val == 'OFF': break
                            count += 1
                        work_blocks[name][day_idx] = count

                team_allowed_zones = {
                    'A': ["L", "T", "T2", "B1", "C1", "B2", "C2", "R", "R1", "R3", "MO", "MO1", "MO2", "S2", "S", "GC", "GB", "GD"],
                    'B': ["T", "T2", "A1", "B1", "C1", "A2", "B2", "C2", "R", "R1", "R2", "R3", "P", "MO", "MO1", "MO2", "S1", "S2", "S", "T2", "GC", "GB", "GD"],
                    'C': ["A1", "B1", "C1", "A2", "B2", "C2", "R", "R1", "R2", "R3", "P", "MO", "MO1", "MO2", "S1", "S2", "S", "GC", "GB", "GD"],
                    'D': ["A1", "B1", "C1", "A2", "B2", "C2", "R1", "R2", "R3", "P", "MO", "MO1", "MO2", "S1", "S2", "GC", "GB", "GD"],
                    'E': ["A1", "C1", "A2", "B2", "R1", "R2", "P", "MO", "MO1", "MO2", "S1", "GB", "GC"],
                    'F': ["A1", "A2", "R2", "MO", "MO1", "P", "S1"],
                    'G': ["A1", "R2", "MO", "MO1", "S1"]
                }

                for day_idx, date_col in enumerate(date_columns):
                    
                    ab_staff = [n for n in all_staff if get_team_of(n, df_shift) in ['A', 'B']]
                    ab_gb_gc_fulfilled = all(sum(monthly_counts[n].get(x, 0) for x in ['GB', 'GC']) >= 1 for n in ab_staff)

                    todays_shifts = df_shift[['姓名', '組別', '性別', date_col]].copy()
                    todays_shifts.columns = ['姓名', '組別', '性別', '班別']
                    todays_template = df_template[['姓名', date_col]].copy()
                    todays_template.columns = ['姓名', '預設區域']
                    daily_data = pd.merge(todays_shifts, todays_template, on='姓名')
                    
                    for shift_type in ['D', 'E', 'N']:
                        shift_staff = daily_data[daily_data['班別'].astype(str).str.upper() == shift_type].copy()
                        if len(shift_staff) == 0: continue
                        
                        unassigned_staff = shift_staff['姓名'].tolist()
                        assignments = {}
                        pre_assigned_zones = []
                        
                        # ==========================================
                        # 第一階段：絕對優先與前置作業
                        # ==========================================
                        # 1️⃣ 拔除手動預排
                        for _, row in shift_staff.iterrows():
                            name, preset = row['姓名'], str(row['預設區域']).strip().upper()
                            if preset not in ['X', 'NAN', 'NONE', ''] and name in unassigned_staff:
                                assignments[name] = preset
                                unassigned_staff.remove(name)
                                if preset not in ['OFF', 'L', 'L2', '職代']:
                                    if get_team_of(name, shift_staff) != 'H': 
                                        pre_assigned_zones.append(preset)
                        
                        # 2️⃣ 拔除當班組長
                        if "L" not in assignments.values():
                            l_chain = [d_l_1, d_l_2, d_l_3] if shift_type == 'D' else [e_l_1, e_l_2, e_l_3] if shift_type == 'E' else [n_l_1, n_l_2, n_l_3]
                            for l_cand in l_chain:
                                if l_cand in unassigned_staff:
                                    assignments[l_cand] = "L"
                                    unassigned_staff.remove(l_cand)
                                    break
                            if "L" not in assignments.values():
                                missing_l_records.append({
                                    '日期': f"2026/06/{str(date_col).zfill(2)}",
                                    '對象班別': f"👑 {shift_type} 班 組長",
                                    '異常提示': "⚠️ 該班別設定的三個順位組長今天全部休假，缺少L，請手動指派"
                                })

                        # 3️⃣ 精算當班常規坑位
                        staff_for_quota = [n for n in unassigned_staff if get_team_of(n, shift_staff) != 'H']
                        total_needed = len(staff_for_quota) + len(pre_assigned_zones)
                        base_zones = (master_zones * 2)[:total_needed] 
                        available_zones = base_zones.copy()
                        for pz in pre_assigned_zones:
                            if pz in available_zones: available_zones.remove(pz)
                            elif available_zones: available_zones.pop()

                        # ==========================================
                        # 第二階段：影子與特權進場
                        # ==========================================
                        # 4️⃣ H 組新人影子帶飛 (反轉邏輯：先定H，再拉老師)
                        for nh_name in h_team_members:
                            if nh_name in unassigned_staff or nh_name in assignments:
                                
                                # A. 決定 H組 去哪區
                                chosen_nh_zone = assignments.get(nh_name, None)
                                allowed_zones = h_team_config.get(nh_name, {}).get("allowed_zones", master_zones)
                                
                                if not chosen_nh_zone:
                                    if day_idx > 0:
                                        prev_shift = str(df_shift.loc[df_shift['姓名'] == nh_name, date_columns[day_idx - 1]].values[0]).strip().upper()
                                        if prev_shift != 'OFF':
                                            prev_zone = str(df_result.loc[df_result['姓名'] == nh_name, date_columns[day_idx - 1]].values[0]).strip()
                                            if prev_zone not in ['OFF', 'L', 'L2', 'X', 'NAN', '', '⚠️需指派老師']:
                                                chosen_nh_zone = prev_zone 
                                    if not chosen_nh_zone:
                                        valid_nh_zones = [z for z in allowed_zones if z in available_zones]
                                        if valid_nh_zones:
                                            valid_nh_zones.sort(key=lambda z: monthly_counts[nh_name].get(z, 0))
                                            chosen_nh_zone = valid_nh_zones[0]
                                        else:
                                            chosen_nh_zone = allowed_zones[0] if allowed_zones else "MO"
                                    
                                    assignments[nh_name] = chosen_nh_zone
                                    if nh_name in unassigned_staff: unassigned_staff.remove(nh_name) 

                                # B. 尋找老師來配對
                                active_preceptor = None
                                for p in h_team_config.get(nh_name, {}).get("preceptors", []):
                                    if p in unassigned_staff:
                                        active_preceptor = p
                                        break
                                
                                if not active_preceptor:
                                    subs = [s for s in unassigned_staff if s != nh_name and get_team_priority(s, shift_staff) <= 5 and get_team_of(s, shift_staff) != 'H' and s not in ["N連啟倫", "N2陳信介"]]
                                    if subs:
                                        subs.sort(key=lambda x: get_team_priority(x, shift_staff))
                                        active_preceptor = subs[0]
                                
                                if not active_preceptor:
                                    subs_any = [s for s in unassigned_staff if s != nh_name and get_team_of(s, shift_staff) != 'H']
                                    if subs_any: active_preceptor = subs_any[0]

                                # C. 指派老師
                                if active_preceptor:
                                    assignments[active_preceptor] = chosen_nh_zone
                                    unassigned_staff.remove(active_preceptor)
                                    if chosen_nh_zone in available_zones: available_zones.remove(chosen_nh_zone)
                                    elif available_zones: available_zones.pop()
                                else:
                                    missing_l_records.append({
                                        '日期': f"2026/06/{str(date_col).zfill(2)}",
                                        '對象班別': f"🌱 新人 {nh_name} ({shift_type}班)",
                                        '異常提示': "⚠️ 設定的專屬教師今天皆未上班（或已去當班別組長），缺乏臨床教師，請人工手動指派"
                                    })
                                    assignments[nh_name] = "⚠️需指派老師"

                        # 5️⃣ 特定人員專屬綁定
                        for special_staff in ["N連啟倫", "N2陳信介"]:
                            if special_staff in unassigned_staff:
                                valid_mo_zones = [z for z in ['MO', 'MO1', 'MO2'] if z in available_zones]
                                if valid_mo_zones:
                                    valid_mo_zones.sort(key=lambda z: monthly_counts[special_staff].get(z, 0))
                                    chosen = valid_mo_zones[0]
                                    assignments[special_staff] = chosen
                                    unassigned_staff.remove(special_staff)
                                    available_zones.remove(chosen)
                                else:
                                    chosen = "MO" if monthly_counts.get(special_staff, {}).get("MO",0) <= monthly_counts.get(special_staff, {}).get("MO1",0) else "MO1"
                                    assignments[special_staff] = chosen
                                    unassigned_staff.remove(special_staff)
                                    if available_zones: available_zones.pop()

                        # ==========================================
                        # 第三階段：連續性與強制連上
                        # ==========================================
                        # 6️⃣ 區塊連續性鎖定 (T/P/G組 絕對連續不中斷)
                        continuous_reqs = []
                        for name in list(unassigned_staff):
                            if day_idx > 0:
                                prev_shift = str(df_shift.loc[df_shift['姓名'] == name, date_columns[day_idx - 1]].values[0]).strip().upper()
                                if prev_shift != 'OFF':
                                    prev_zone = str(df_result.loc[df_result['姓名'] == name, date_columns[day_idx - 1]].values[0]).strip()
                                    if prev_zone not in ['OFF', 'L', 'X', 'NAN', '']:
                                        is_train = (name in train_s2 and prev_zone == "S2") or (name in train_b1_r and prev_zone in ["B1", "R", "S", "C2"])
                                        is_TP = prev_zone in ["T", "T2", "P", "P2"]
                                        is_team_G = (get_team_of(name, shift_staff) == 'G')
                                        
                                        if is_train or is_TP or is_team_G:
                                            req_zones = [prev_zone]
                                            if is_TP:
                                                req_zones = ['T', 'T2'] if prev_zone in ['T', 'T2'] else ['P', 'P2']
                                            continuous_reqs.append({'name': name, 'zones': req_zones, 'is_train': is_train})
                        
                        continuous_reqs.sort(key=lambda x: not x['is_train']) 
                        for req in continuous_reqs:
                            name, z_list = req['name'], req['zones']
                            if name in unassigned_staff:
                                valid_z = [z for z in z_list if z in available_zones]
                                if valid_z:
                                    valid_z.sort(key=lambda z: monthly_counts[name].get(z, 0))
                                    chosen_z = valid_z[0]
                                    available_zones.remove(chosen_z)
                                else:
                                    # ⚠️ 強制保護：就算坑位沒了，也硬生生塞進去，保證 OFF 前不斷鏈
                                    chosen_z = z_list[0]
                                    if available_zones: available_zones.pop() 
                                assignments[name] = chosen_z
                                unassigned_staff.remove(name)

                        # 7️⃣ 白名單剛放假回來的優先進場
                        for t_list, t_zone in [(train_s2, "S2"), (train_t, "T")]:
                            cands = [n for n in list(unassigned_staff) if n in t_list]
                            if cands and t_zone in available_zones:
                                cands.sort(key=lambda n: monthly_counts[n].get(t_zone, 0))
                                assignments[cands[0]] = t_zone
                                unassigned_staff.remove(cands[0])
                                available_zones.remove(t_zone)

                        # ==========================================
                        # 第四階段：戰略護盤與一般分配
                        # ==========================================
                        # 8️⃣ 內、外、急大區主力護盤 (1位A 或 1位B)
                        macro_target_zones = {
                            'MED': ['A2', 'B2', 'C2'],
                            'SURG': ['S', 'S1', 'S2'],
                            'RESUS': ['R', 'R1', 'R3']
                        }
                        for macro, z_list in macro_target_zones.items():
                            a_count = sum(1 for n, z in assignments.items() if get_team_of(n, shift_staff) == 'A' and z in z_list)
                            if a_count >= 1: continue 
                                
                            assigned_macro = False
                            avail_A = [n for n in unassigned_staff if get_team_of(n, shift_staff) == 'A']
                            if avail_A:
                                best_a, best_z, best_score = None, None, float('inf')
                                for a_cand in avail_A:
                                    valid_mz = [z for z in z_list if z in available_zones]
                                    valid_mz = [z for z in valid_mz if z not in ['R2', 'A2', 'S1']]
                                    gender_a = shift_staff[shift_staff['姓名'] == a_cand]['性別'].values[0] if len(shift_staff[shift_staff['姓名'] == a_cand]) > 0 else ""
                                    valid_mz = [z for z in valid_mz if not (z == "S2" and str(gender_a).upper() == "M")]
                                    for z in valid_mz:
                                        score = get_zone_score(z, a_cand, 'A', day_idx, df_result, date_columns, monthly_counts, work_blocks, {}, ab_gb_gc_fulfilled)
                                        if score < best_score:
                                            best_score, best_a, best_z = score, a_cand, z
                                if best_a and best_score < 4000000:
                                    assignments[best_a] = best_z
                                    unassigned_staff.remove(best_a)
                                    available_zones.remove(best_z)
                                    assigned_macro = True

                            if not assigned_macro:
                                avail_B = [n for n in unassigned_staff if get_team_of(n, shift_staff) == 'B']
                                if avail_B:
                                    best_b, best_z_b, best_score_b = None, None, float('inf')
                                    for b_cand in avail_B:
                                        valid_mz = [z for z in z_list if z in available_zones]
                                        gender_b = shift_staff[shift_staff['姓名'] == b_cand]['性別'].values[0] if len(shift_staff[shift_staff['姓名'] == b_cand]) > 0 else ""
                                        valid_mz = [z for z in valid_mz if not (z == "S2" and str(gender_b).upper() == "M")]
                                        for z in valid_mz:
                                            score = get_zone_score(z, b_cand, 'B', day_idx, df_result, date_columns, monthly_counts, work_blocks, {}, ab_gb_gc_fulfilled)
                                            if score < best_score_b:
                                                best_score_b, best_b, best_z_b = score, b_cand, z
                                    if best_b and best_score_b < 4000000:
                                        assignments[best_b] = best_z_b
                                        unassigned_staff.remove(best_b)
                                        available_zones.remove(best_z_b)

                        # 9️⃣ 剩下的常規人員「貪婪分配」
                        current_day_macro_teams = {}
                        for assigned_name, assigned_z in assignments.items():
                            t_val = get_team_of(assigned_name, shift_staff)
                            m_val = get_macro(assigned_z)
                            if m_val not in current_day_macro_teams: current_day_macro_teams[m_val] = set()
                            current_day_macro_teams[m_val].add(t_val)

                        random.shuffle(unassigned_staff)
                        unassigned_staff.sort(key=lambda x: get_team_priority(x, shift_staff))

                        for name in list(unassigned_staff):
                            gender = shift_staff[shift_staff['姓名'] == name]['性別'].values[0] if len(shift_staff[shift_staff['姓名'] == name]) > 0 else ""
                            team = get_team_of(name, shift_staff)
                            team_allowed = team_allowed_zones.get(team, available_zones)

                            valid_cands = [z for z in available_zones if z in team_allowed and not (z == "S2" and str(gender).upper() == "M")]
                            if not valid_cands: 
                                valid_cands = [z for z in available_zones if not (z == "S2" and str(gender).upper() == "M")]
                                if team == 'A': valid_cands = [z for z in valid_cands if z not in ['R2', 'A2', 'S1']]
                                if not valid_cands: valid_cands = available_zones.copy()

                            z_scores = [(z, get_zone_score(z, name, team, day_idx, df_result, date_columns, monthly_counts, work_blocks, current_day_macro_teams, ab_gb_gc_fulfilled)) for z in valid_cands]
                            z_scores.sort(key=lambda x: x[1])
                            
                            chosen_zone = z_scores[0][0]
                            assignments[name] = chosen_zone
                            unassigned_staff.remove(name)
                            if chosen_zone in available_zones: available_zones.remove(chosen_zone)
                            
                            m_val = get_macro(chosen_zone)
                            if m_val not in current_day_macro_teams: current_day_macro_teams[m_val] = set()
                            current_day_macro_teams[m_val].add(team)

                        for name, assigned_zone in assignments.items():
                            df_result.loc[df_result['姓名'] == name, date_col] = assigned_zone
                            monthly_counts[name][assigned_zone] = monthly_counts[name].get(assigned_zone, 0) + 1

                    progress_bar.progress((day_idx + 1) / len(date_columns))

                name_col_index = df_result.columns.get_loc('姓名')
                majority_shift_dict = {name: max(set([x for x in df_shift[df_shift['姓名'] == name][date_columns].values.flatten() if x in ['D', 'E', 'N']]), key=[x for x in df_shift[df_shift['姓名'] == name][date_columns].values.flatten() if x in ['D', 'E', 'N']].count) if [x for x in df_shift[df_shift['姓名'] == name][date_columns].values.flatten() if x in ['D', 'E', 'N']] else "" for name in all_staff}
                df_result.insert(name_col_index, '當月班別', df_result['姓名'].map(majority_shift_dict))
                
                # 🌟 V23: 更新欄位名稱與順序
                summary_cols = ['2區計', 'MO計', 'R1+R3計', 'GB+GC計']
                zone_count_order = ["L", "L2", "T", "T2", "GB", "GC", "GD", "A1", "B1", "C1", "A2", "B2", "C2", "R", "R1", "R2", "R3", "S1", "S2", "S", "P", "P2", "MO", "MO1", "MO2", "職代"]
                
                df_excel = df_result.copy()

                for count_zone in zone_count_order:
                    df_result[count_zone] = df_result[date_columns].apply(lambda row: (row == count_zone).sum(), axis=1)
                
                df_result['2區計'] = df_result[['A2', 'B2', 'C2']].sum(axis=1) if all(x in df_result.columns for x in ['A2', 'B2', 'C2']) else 0
                df_result['MO計'] = df_result[['MO', 'MO1', 'MO2']].sum(axis=1) if all(x in df_result.columns for x in ['MO', 'MO1', 'MO2']) else 0
                df_result['R1+R3計'] = df_result[['R1', 'R3']].sum(axis=1) if all(x in df_result.columns for x in ['R1', 'R3']) else 0
                df_result['GB+GC計'] = df_result[['GB', 'GC']].sum(axis=1) if all(x in df_result.columns for x in ['GB', 'GC']) else 0
                df_result = df_result.fillna("")

                first_date_col_idx = df_excel.columns.get_loc(date_columns[0]) + 1
                last_date_col_idx = df_excel.columns.get_loc(date_columns[-1]) + 1
                col_start = get_column_letter(first_date_col_idx)
                col_end = get_column_letter(last_date_col_idx)

                # 將 summary_cols 放在前面
                target_cols = summary_cols + zone_count_order
                for col in target_cols:
                    if col not in df_excel.columns:
                        df_excel[col] = ""

                for i in range(len(df_excel)):
                    row_excel = i + 2 
                    range_str = f"{col_start}{row_excel}:{col_end}{row_excel}"
                    for z in zone_count_order: df_excel.at[i, z] = f'=COUNTIF({range_str}, "{z}")'
                    df_excel.at[i, '2區計'] = f'=COUNTIF({range_str}, "A2")+COUNTIF({range_str}, "B2")+COUNTIF({range_str}, "C2")'
                    df_excel.at[i, 'MO計'] = f'=COUNTIF({range_str}, "MO")+COUNTIF({range_str}, "MO1")+COUNTIF({range_str}, "MO2")'
                    df_excel.at[i, 'R1+R3計'] = f'=COUNTIF({range_str}, "R1")+COUNTIF({range_str}, "R3")'
                    df_excel.at[i, 'GB+GC計'] = f'=COUNTIF({range_str}, "GB")+COUNTIF({range_str}, "GC")'

                if len(missing_l_records) == 0:
                    df_missing_l = pd.DataFrame([{
                        '日期': '🎉 完美平衡！',
                        '對象班別': '全月皆無異常',
                        '異常提示': '本月天天都有組長與臨床教師，無任何缺乏提示，不需進行人工修改。'
                    }])
                else:
                    df_missing_l = pd.DataFrame(missing_l_records)

                st.success("🎉 排班完成！已更新統計結果呈現順序與欄位名稱！")
                
                def preview_highlight(val):
                    try:
                        v = pd.to_numeric(val)
                        if v == 0: return 'background-color: #FFFF00; color: #000000'
                        elif v > 5: return 'background-color: #FFDAB9; color: #000000'
                    except: pass
                    return ''
                
                target_cols_preview = [c for c in target_cols if c in preview_df.columns]
                st.dataframe(preview_df.style.map(preview_highlight, subset=target_cols_preview))

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_excel.to_excel(writer, index=False, sheet_name='排班結果')
                    df_shift.to_excel(writer, index=False, sheet_name='原始班表')
                    df_missing_l.to_excel(writer, index=False, sheet_name='缺乏組長與教師提示') 
                    
                    ws = writer.sheets['排班結果']
                    
                    start_summary_col_idx = df_excel.columns.get_loc(target_cols[0]) + 1
                    end_summary_col_idx = len(df_excel.columns)
                    start_sum_col_letter = get_column_letter(start_summary_col_idx)
                    end_sum_col_letter = get_column_letter(end_summary_col_idx)
                    format_range = f"{start_sum_col_letter}2:{end_sum_col_letter}{len(df_excel)+1}"
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    peach_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
                    ws.conditional_formatting.add(format_range, CellIsRule(operator='equal', formula=['0'], fill=yellow_fill))
                    ws.conditional_formatting.add(format_range, CellIsRule(operator='greaterThan', formula=['5'], fill=peach_fill))

                st.download_button("📥 下載排班表", data=output.getvalue(), file_name="初步自動生成排班結果.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error("❌ **系統運算失敗！** 請參考右側的「💡常見錯誤排除指南」檢查您的檔案或側邊欄設定。")
                st.warning(f"🔧 開發人員除錯代碼 (供參考)：\n\n {str(e)}")
